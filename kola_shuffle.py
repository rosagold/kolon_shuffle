#!/usr/bin/env python

import sys
import random
import openpyxl
import logging

__author__ = "rosagold"
__email__ = "meinschallundrauch@gmail.com"
__copyright__ = "Me and only me and everyone else too"
__version__ = "0.1.0"

usage = f"""
Usage: ./kola_shuffle.py FILE [NR] [OFFSET]
Shuffle rows by column in an excel spreadsheet.
    
    FILE    input file (.xlsx) to shuffle
    NR      desired number of results (default 1)
    SKIP    header rows to skip (default 2)

version: {__version__}
"""

if __name__ == "__main__":
    args = sys.argv[1:]
    path, n, offset = None, 1, 2

    if "--debug" in args:
        logging.basicConfig(
            level="DEBUG", format="[%(asctime)s] %(levelname)s: %(message)s"
        )
        logging.debug(f"args: {args}")
        args.remove("--debug")

    logging.debug(f"version: {__version__}")

    if len(args) == 1:
        path = args[0]
    elif len(args) == 2:
        path, n = args
    elif len(args) == 3:
        path, n, offset = args
    else:
        print("Wrong number of arguments." + usage)
        exit(1)

    n = int(n)
    offset = int(offset)
    wb = openpyxl.load_workbook(filename=path, read_only=True)

    # read file and use first sheet
    sheet = wb.sheetnames[0]
    logging.debug(f"sheet: {sheet}")
    sheet = wb[sheet]

    if offset > sheet.max_row:
        print(
            "Cannot skip more (header) rows than the file contains. "
            "Either fill document or adjust SKIP parameter."
        )
        exit(1)

    logging.debug(f"random state: {random.getstate()}")

    baselist = list(range(offset + 1, sheet.max_row + 1))
    sz = len(baselist)
    for _ in range(n):
        s = ""
        for c in range(1, sheet.max_column + 1):
            sample = random.sample(baselist, sz)
            for i, r in enumerate(sample):
                v = sheet.cell(r, c).value
                logging.debug(f"cell({r}, {c}) = {v} (random try nr {i+1})")
                if v is not None:
                    break
            else:
                logging.warning("No valid value found - skipping column")
                continue
            s += f"{v} "
        print(s)
